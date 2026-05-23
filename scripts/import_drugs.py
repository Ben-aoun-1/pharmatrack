"""
PharmTrack — Drug DB import.

Reads a drugs spreadsheet (.xls or .xlsx), pre-computes each selling price
(CLAUDE.md section 6), then:
  1. Upserts every drug into the Supabase `drugs` table (service-role client).
  2. Exports a self-contained SQLite file for distribution to Windows agents.

The selling price is computed ONCE here, at import time, and stored. It is
never recalculated at runtime — neither the dashboard nor the agent recomputes
it (CLAUDE.md sections 6 and 12).

Usage:
    python import_drugs.py drugs.xlsx
    python import_drugs.py drugs.xlsx --sqlite-out drugs.sqlite --sheet 0

Required environment variables (see .env.example):
    NEXT_PUBLIC_SUPABASE_URL
    SUPABASE_SERVICE_ROLE_KEY        # server-only, never expose to the browser
"""

import argparse
import os
import sqlite3
import sys

from dotenv import load_dotenv
from openpyxl import load_workbook
from supabase import create_client

# ---------------------------------------------------------------------------
# Domain logic — drug pricing (mirror of CLAUDE.md section 6).
# This is the Python source of truth for margins. The TypeScript copy lives in
# lib/constants.ts and the agent copy in agent/config.py. One source per
# language — keep them identical.
# ---------------------------------------------------------------------------
MARGIN_BY_CATEGORY = {
    "REMBOURSABLE": 0.28,
    "OTC":          0.33,
    "GENERIQUE":    0.30,
}
DEFAULT_MARGIN = 0.30


def compute_selling_price(tarif_reference: float, categorie: str) -> float:
    margin = MARGIN_BY_CATEGORY.get(categorie.upper().strip(), DEFAULT_MARGIN)
    return round(tarif_reference * (1 + margin), 3)


# ---------------------------------------------------------------------------
# Spreadsheet column mapping.
# Header names are matched case-insensitively against the first row.
# ---------------------------------------------------------------------------
COLUMN_ALIASES = {
    "code_pct":        ("CODE_PCT", "CODE PCT", "CODEPCT", "BARCODE"),
    "nom_commercial":  ("NOM_COMMERCIAL", "NOM COMMERCIAL", "NOM"),
    "tarif_reference": ("TARIF_REFERENCE", "TARIF REFERENCE", "TARIF", "PRIX"),
    "categorie":       ("CATEGORIE", "CATÉGORIE", "CATEGORY"),
    "ap":              ("AP",),
}


def resolve_columns(header_row) -> dict:
    """Map our field names to 0-based column indices from the header row."""
    # Normalize each header cell to an upper-cased, trimmed string.
    headers = {}
    for idx, cell in enumerate(header_row):
        if cell is None:
            continue
        headers[str(cell).strip().upper()] = idx

    mapping = {}
    for field, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            if alias.upper() in headers:
                mapping[field] = headers[alias.upper()]
                break

    missing = [f for f in ("code_pct", "nom_commercial", "tarif_reference")
               if f not in mapping]
    if missing:
        raise ValueError(
            f"Spreadsheet is missing required column(s): {', '.join(missing)}. "
            f"Found headers: {', '.join(sorted(headers))}"
        )
    return mapping


def _iter_rows(path: str, sheet):
    """Yield rows (sequences of cell values), choosing the engine by extension.

    .xlsx/.xlsm → openpyxl (engine for modern Excel).
    .xls        → xlrd (the only engine that still reads the legacy format).
    """
    ext = os.path.splitext(path)[1].lower()

    if ext == ".xls":
        # openpyxl cannot read the legacy .xls format; use xlrd instead.
        import xlrd  # imported lazily so .xlsx-only setups don't need it

        book = xlrd.open_workbook(path)
        worksheet = (
            book.sheet_by_index(sheet)
            if isinstance(sheet, int)
            else book.sheet_by_name(sheet)
        )
        for r in range(worksheet.nrows):
            # Normalize xlrd's empty cells ("") to None to match openpyxl.
            yield [
                None if (isinstance(v, str) and v.strip() == "") else v
                for v in worksheet.row_values(r)
            ]
        return

    workbook = load_workbook(filename=path, read_only=True, data_only=True)
    worksheet = (
        workbook.worksheets[sheet] if isinstance(sheet, int) else workbook[sheet]
    )
    try:
        for row in worksheet.iter_rows(values_only=True):
            yield row
    finally:
        workbook.close()


def read_rows(path: str, sheet) -> list:
    """Read the spreadsheet into a list of drug dicts with computed prices."""
    rows = _iter_rows(path, sheet)
    try:
        header_row = next(rows)
    except StopIteration:
        raise ValueError("Spreadsheet is empty.")

    cols = resolve_columns(header_row)

    def value(row, field):
        idx = cols.get(field)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    drugs = []
    skipped = 0
    for row in rows:
        code_raw = value(row, "code_pct")
        tarif_raw = value(row, "tarif_reference")
        nom = value(row, "nom_commercial")

        # A drug needs a barcode, a name, and a reference price to be usable.
        if code_raw is None or nom is None or tarif_raw is None:
            skipped += 1
            continue

        # CODE_PCT is a 7-digit barcode; Excel may have read it as a number.
        code_pct = str(code_raw).strip()
        if code_pct.endswith(".0"):
            code_pct = code_pct[:-2]

        categorie = "" if value(row, "categorie") is None \
            else str(value(row, "categorie")).strip()
        ap = None if value(row, "ap") is None else str(value(row, "ap")).strip()

        try:
            tarif_reference = float(tarif_raw)
        except (TypeError, ValueError):
            skipped += 1
            continue

        drugs.append({
            "code_pct":        code_pct,
            "nom_commercial":  str(nom).strip(),
            "tarif_reference": tarif_reference,
            "selling_price":   compute_selling_price(tarif_reference, categorie),
            "categorie":       categorie or None,
            "ap":              ap,
        })

    if skipped:
        print(f"  ⚠ Skipped {skipped} row(s) with missing/invalid required fields.")
    return drugs


def upsert_to_supabase(drugs: list, batch_size: int = 500) -> None:
    """Upsert all drugs into the Supabase `drugs` table via the service role."""
    load_dotenv()
    url = os.environ.get("NEXT_PUBLIC_SUPABASE_URL")
    key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    if not url or not key:
        print(
            "ERROR: NEXT_PUBLIC_SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY must "
            "be set (see .env.example).",
            file=sys.stderr,
        )
        sys.exit(1)

    client = create_client(url, key)

    total = len(drugs)
    for start in range(0, total, batch_size):
        batch = drugs[start:start + batch_size]
        # on_conflict=code_pct → idempotent re-imports update existing rows.
        client.table("drugs").upsert(batch, on_conflict="code_pct").execute()
        print(f"  ↳ Supabase upserted {min(start + batch_size, total)}/{total}")


def export_sqlite(drugs: list, sqlite_path: str) -> None:
    """Write a fresh SQLite file the Windows agent uses for offline lookup."""
    # Rebuild from scratch so the exported DB always matches the spreadsheet.
    if os.path.exists(sqlite_path):
        os.remove(sqlite_path)

    conn = sqlite3.connect(sqlite_path)
    try:
        conn.execute(
            """
            CREATE TABLE drugs (
                code_pct        TEXT PRIMARY KEY,
                nom_commercial  TEXT NOT NULL,
                tarif_reference REAL,
                selling_price   REAL NOT NULL,
                categorie       TEXT,
                ap              TEXT
            )
            """
        )
        conn.executemany(
            """
            INSERT INTO drugs
                (code_pct, nom_commercial, tarif_reference, selling_price,
                 categorie, ap)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            [
                (
                    d["code_pct"],
                    d["nom_commercial"],
                    d["tarif_reference"],
                    d["selling_price"],
                    d["categorie"],
                    d["ap"],
                )
                for d in drugs
            ],
        )
        conn.commit()
    finally:
        conn.close()
    print(f"  ↳ SQLite exported → {sqlite_path}")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Import a drugs spreadsheet (.xls/.xlsx) into Supabase and "
        "export a SQLite file."
    )
    parser.add_argument(
        "xlsx", help="Path to the source spreadsheet (.xls or .xlsx)."
    )
    parser.add_argument(
        "--sqlite-out",
        default="drugs.sqlite",
        help="Output path for the exported SQLite file (default: drugs.sqlite).",
    )
    parser.add_argument(
        "--sheet",
        default=0,
        help="Worksheet index (int) or name to read (default: first sheet).",
    )
    parser.add_argument(
        "--no-upload",
        action="store_true",
        help="Skip the Supabase upsert; only export the SQLite file.",
    )
    args = parser.parse_args()

    if not os.path.exists(args.xlsx):
        print(f"ERROR: file not found: {args.xlsx}", file=sys.stderr)
        sys.exit(1)

    # --sheet may be an int index or a sheet name.
    sheet = args.sheet
    if isinstance(sheet, str) and sheet.isdigit():
        sheet = int(sheet)

    print(f"Reading {args.xlsx} ...")
    drugs = read_rows(args.xlsx, sheet)
    print(f"  ↳ Parsed {len(drugs)} drug(s).")
    if not drugs:
        print("Nothing to import.", file=sys.stderr)
        sys.exit(1)

    if args.no_upload:
        print("Skipping Supabase upsert (--no-upload).")
    else:
        print("Upserting into Supabase ...")
        upsert_to_supabase(drugs)

    print("Exporting SQLite ...")
    export_sqlite(drugs, args.sqlite_out)

    print("Done.")


if __name__ == "__main__":
    main()
